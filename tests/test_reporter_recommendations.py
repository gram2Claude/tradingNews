"""Tests for reporter recommendations behaviour (task 06 T5 + task 08 δ-completion).

Covers:
* recommendations from recommendations table → routed to recommendations/ folder
* deterministic numbering when timestamps collide (tie-break)
* NULL structural fields в frontmatter отсутствуют (не "None")
* xlsx второй лист содержит структурные колонки
* persons.csv агрегирует упоминания из обеих junction-таблиц

После δ-completion (task 08): news table не имеет item_type — рекомендации
живут только в recommendations table. Legacy «finam через news.item_type=
'recommendation'» путь убран; тесты упрощены.
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import db, reporter
from src.config import Config, CompanyCfg, SourceCfg


@pytest.fixture
def cfg(tmp_path: Path) -> Config:
    seed = tmp_path / "seed.csv"
    seed.write_text(
        "full_name,status,brand\nИгорь Шехтерман,CEO,X5\n",
        encoding="utf-8",
    )
    return Config(
        start_date="2026-05-01",
        llm_provider="openai",
        llm_model="gpt-5-mini",
        output_root=tmp_path / "out",
        db_path=tmp_path / "db.sqlite",
        auto_run=False,
        timezone="Europe/Moscow",
        companies=[CompanyCfg("X5", None, ["x5_ir"], str(seed))],
        sources={"x5_ir": SourceCfg("x5_ir", "X5 IR", "https://www.x5.ru/", "x5_ir", True)},
    )


def _insert_recommendation(
    conn: sqlite3.Connection,
    *,
    url: str,
    headline: str,
    body: str,
    published_utc: datetime,
    mood: str = "neutral",
    mood_reason: str = "ok",
    source_code: str = "x5_ir",
    target_price: float | None = None,
    recommendation_action: str | None = None,
    potential_pct: float | None = None,
    multipliers_json: str | None = None,
) -> int:
    cid = conn.execute("SELECT id FROM companies WHERE name='X5'").fetchone()["id"]
    sid = conn.execute("SELECT id FROM sources WHERE code=?", (source_code,)).fetchone()["id"]
    cur = conn.execute(
        "INSERT INTO recommendations (company_id, source_id, url, headline, body, "
        "published_at, mood, mood_reason, target_price, recommendation_action, "
        "potential_pct, multipliers_json, status) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'analyzed')",
        (
            cid, sid, url, headline, body, published_utc.isoformat(),
            mood, mood_reason, target_price, recommendation_action,
            potential_pct, multipliers_json,
        ),
    )
    return cur.lastrowid


def _insert_news(
    conn: sqlite3.Connection,
    *,
    url: str,
    headline: str,
    body: str,
    published_utc: datetime,
    mood: str = "neutral",
    source_code: str = "x5_ir",
) -> int:
    """δ-completion: news не имеет item_type. Helper только для plain news."""
    cid = conn.execute("SELECT id FROM companies WHERE name='X5'").fetchone()["id"]
    sid = conn.execute("SELECT id FROM sources WHERE code=?", (source_code,)).fetchone()["id"]
    cur = conn.execute(
        "INSERT INTO news (company_id, source_id, url, headline, body, "
        "published_at, mood, mood_reason, status) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, 'reason', 'analyzed')",
        (cid, sid, url, headline, body, published_utc.isoformat(), mood),
    )
    return cur.lastrowid


# ----------------------------------------------------------- routing


def test_recommendation_routes_to_recommendations_folder(cfg: Config) -> None:
    """Row из таблицы recommendations попадает в output/X5/recommendations/."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_recommendation(
        conn,
        url="https://lmsic.com/x5",
        headline="X5 разбор",
        body="Body",
        published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
        target_price=3200.0,
        recommendation_action="hold",
        potential_pct=12.5,
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")

    news_dir = cfg.output_root / "X5" / "news"
    rec_dir = cfg.output_root / "X5" / "recommendations"
    news_files = list(news_dir.rglob("*.md")) if news_dir.exists() else []
    rec_files = list(rec_dir.rglob("*.md"))
    assert news_files == []
    assert len(rec_files) == 1


def test_finam_classified_rec_routes_to_recommendations(cfg: Config) -> None:
    """δ-completion: после per-item dispatch finam-rec уже в recommendations
    table (analyzer переместил). Reporter читает оттуда и кладёт в
    output/X5/recommendations/."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_recommendation(
        conn, url="https://finam.ru/r1", headline="finam rec post-dispatch",
        body="Body", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
        # NULL structural fields — типичная finam-row после per-item dispatch
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")

    rec_files = list((cfg.output_root / "X5" / "recommendations").rglob("*.md"))
    assert len(rec_files) == 1


# ----------------------------------------------------------- tie-break


def test_recommendations_tie_break_stable_numbering(cfg: Config) -> None:
    """Два recommendation item'а с одинаковым published_at — deterministic
    order по (source_code ASC, url ASC). Повторный report_all даёт идентичные
    имена файлов (idempotency)."""
    db.init_db(cfg)
    # Регистрируем дополнительный source чтобы можно было различать по source_code
    conn = db.connect(cfg.db_path)
    conn.execute(
        "INSERT INTO sources (code, name, base_url, enabled) "
        "VALUES ('lmsic', 'LMS', 'https://lmsic.com/', 1)"
    )
    same_ts = datetime(2026, 5, 19, 10, 0, tzinfo=timezone.utc)
    _insert_recommendation(
        conn, url="https://finam.ru/r1", headline="finam recommendation",
        body="Body", published_utc=same_ts, source_code="x5_ir",  # уже из cfg fixture
    )
    _insert_recommendation(
        conn, url="https://lmsic.com/r1", headline="lmsic recommendation",
        body="Body", published_utc=same_ts, target_price=100.0,
        source_code="lmsic",
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    rec_files1 = {p.name: p.read_text(encoding="utf-8")
                  for p in (cfg.output_root / "X5" / "recommendations").rglob("*.md")}

    reporter.report_all(cfg, "X5")
    rec_files2 = {p.name: p.read_text(encoding="utf-8")
                  for p in (cfg.output_root / "X5" / "recommendations").rglob("*.md")}

    assert set(rec_files1.keys()) == set(rec_files2.keys())
    assert len(rec_files1) == 2

    # ORDER BY source_code ASC: 'lmsic' < 'x5_ir' → lmsic gets _01
    file_01 = next(name for name in rec_files1 if "_01.md" in name)
    file_02 = next(name for name in rec_files1 if "_02.md" in name)
    assert "lmsic recommendation" in rec_files1[file_01]
    assert "finam recommendation" in rec_files1[file_02]


# ----------------------------------------------------------- frontmatter NULL


def test_recommendation_frontmatter_omits_null_structural_fields(cfg: Config) -> None:
    """target_price=NULL → ключ target_price отсутствует во frontmatter."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_recommendation(
        conn, url="https://lmsic.com/r", headline="без структуры",
        body="Body", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
        # все extras None
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    md = next((cfg.output_root / "X5" / "recommendations").rglob("*.md"))
    text = md.read_text(encoding="utf-8")
    assert "target_price" not in text
    assert "recommendation_action" not in text
    assert "potential_pct" not in text
    assert "multipliers_json" not in text


def test_recommendation_frontmatter_includes_non_null_extras(cfg: Config) -> None:
    """Заполненные структурные поля попадают во frontmatter."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_recommendation(
        conn, url="https://lmsic.com/r", headline="с target",
        body="Body", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
        target_price=3200.5,
        recommendation_action="hold",
        potential_pct=12.5,
        multipliers_json='{"P/E":6.8}',
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    md = next((cfg.output_root / "X5" / "recommendations").rglob("*.md"))
    text = md.read_text(encoding="utf-8")
    assert "target_price: 3200.5" in text
    assert "recommendation_action: hold" in text
    assert "potential_pct: 12.5" in text
    assert 'multipliers_json:' in text  # JSON-строка экранируется


# ----------------------------------------------------------- xlsx two sheets


def test_xlsx_recommendations_sheet_has_structural_columns(cfg: Config) -> None:
    """Второй лист 'recommendations' содержит target_price, action, potential, multipliers."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_recommendation(
        conn, url="https://lmsic.com/r", headline="X5 hold",
        body="Body", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
        target_price=3200.0, recommendation_action="hold", potential_pct=12.5,
        multipliers_json='{"P/E":6.8}',
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    wb = load_workbook(cfg.output_root / "X5" / "news_list" / "data.xlsx")
    assert wb.sheetnames == ["news", "recommendations"]

    ws = wb["recommendations"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers == [
        "date", "headline", "persons", "mood", "source",
        "target_price", "recommendation_action", "potential_pct", "multipliers",
    ]
    # Data row
    assert ws.cell(2, 2).value == "X5 hold"
    assert ws.cell(2, 6).value == 3200.0
    assert ws.cell(2, 7).value == "hold"
    assert ws.cell(2, 8).value == 12.5
    assert ws.cell(2, 9).value == '{"P/E":6.8}'


# ----------------------------------------------------------- persons union


def test_persons_csv_unions_news_and_recommendation_mentions(cfg: Config) -> None:
    """persons.csv считает упоминания person'а в news_persons И recommendation_persons."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    # Person Шехтерман уже из seed
    person_id = conn.execute(
        "SELECT id FROM persons WHERE full_name='Игорь Шехтерман'"
    ).fetchone()["id"]

    # Один news + linked person (item_type убран — все news = news после δ)
    nid = _insert_news(
        conn, url="https://x.x/n1", headline="X5 news",
        body="Шехтерман прокомментировал", mood="pos",
        published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
    )
    conn.execute("INSERT INTO news_persons (news_id, person_id) VALUES (?, ?)", (nid, person_id))

    # Одна recommendation + linked person с другим mood
    rid = _insert_recommendation(
        conn, url="https://lmsic.com/r1", headline="X5 rec",
        body="Шехтерман в рекомендации", mood="neg",
        published_utc=datetime(2026, 5, 20, tzinfo=timezone.utc),
    )
    conn.execute(
        "INSERT INTO recommendation_persons (recommendation_id, person_id) VALUES (?, ?)",
        (rid, person_id),
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    csv_text = (cfg.output_root / "X5" / "affiliate" / "persons.csv").read_text(encoding="utf-8")
    lines = csv_text.strip().split("\n")
    # Header + 1 row (Шехтерман — единственный person в seed)
    assert len(lines) == 2
    # row format: person, status, brand, pos_freq, neg_freq, zero_freq, total_freq
    parts = lines[1].split(",")
    assert parts[0] == "Игорь Шехтерман"
    pos_freq, neg_freq, zero_freq, total_freq = parts[3], parts[4], parts[5], parts[6]
    assert pos_freq == "1"  # news pos
    assert neg_freq == "1"  # rec neg
    assert zero_freq == "0"
    assert total_freq == "2"  # обе таблицы суммируются
