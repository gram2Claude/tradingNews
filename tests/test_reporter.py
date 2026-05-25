"""Tests for the reporter: slugs, frontmatter, file layout, idempotency."""
from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src import db, reporter
from src.config import Config, CompanyCfg, SourceCfg


@pytest.fixture
def cfg(tmp_path: Path) -> Config:
    seed = tmp_path / "seed.csv"
    seed.write_text(
        "full_name,status,brand\n"
        "Игорь Шехтерман,CEO,X5 Retail Group\n"
        "Ольга Наумова,руководитель бренда,Пятёрочка\n",
        encoding="utf-8",
    )
    return Config(
        start_date="2026-05-01",
        llm_provider="openai",
        llm_model="gpt-5-mini",
        output_root=tmp_path / "out",
        db_path=tmp_path / "db.sqlite",
        auto_run=False,
        timezone="Europe/Moscow",
        companies=[CompanyCfg("X5", None, ["x5_ir"], str(seed))],
        sources={"x5_ir": SourceCfg("x5_ir", "X5 IR", "https://www.x5.ru/", "x5_ir", True)},
        openai_api_key=None,
    )


def _insert_news(conn, *, headline, body, mood, published_utc, persons_full_names=(),
                 item_type="news", mood_reason=None):
    cid = conn.execute("SELECT id FROM companies WHERE name='X5'").fetchone()["id"]
    sid = conn.execute("SELECT id FROM sources WHERE code='x5_ir'").fetchone()["id"]
    reason = mood_reason if mood_reason is not None else f"reason for {mood}"
    cur = conn.execute(
        "INSERT INTO news (company_id, source_id, url, headline, body, "
        "published_at, mood, mood_reason, item_type, status) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (cid, sid, f"https://x.x/{headline[:30]}", headline, body,
         published_utc.isoformat(), mood, reason, item_type, "analyzed"),
    )
    nid = cur.lastrowid
    for name in persons_full_names:
        pid = conn.execute(
            "SELECT id FROM persons WHERE full_name = ?", (name,)
        ).fetchone()["id"]
        conn.execute(
            "INSERT INTO news_persons (news_id, person_id) VALUES (?, ?)",
            (nid, pid),
        )
    return nid


def test_report_writes_recommendation_to_separate_folder(cfg: Config) -> None:
    """item_type='recommendation' → output/X5/recommendations/<YYYY>/<YYYY_MM>/*.md;
    item_type='news' → output/X5/news/<YYYY>/<YYYY_MM>/*.md."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(conn, headline="Корп. событие X5", body="Body",
                 mood="pos", published_utc=datetime(2026, 5, 19, 7, tzinfo=timezone.utc),
                 item_type="news")
    _insert_news(conn, headline="Аналитик рекомендует buy X5", body="Body",
                 mood="pos", published_utc=datetime(2026, 5, 20, 7, tzinfo=timezone.utc),
                 item_type="recommendation")
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")

    out = cfg.output_root / "X5"
    news_files = list((out / "news").rglob("*.md"))
    rec_files = list((out / "recommendations").rglob("*.md"))
    assert len(news_files) == 1
    assert len(rec_files) == 1
    # Each MD has the item_type in frontmatter
    assert "item_type: news" in news_files[0].read_text(encoding="utf-8")
    assert "item_type: recommendation" in rec_files[0].read_text(encoding="utf-8")


def test_report_does_not_create_empty_recommendations_folder(cfg: Config) -> None:
    """Если 0 recommendations — папка recommendations/ не должна создаваться (codex P2.3)."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(conn, headline="Only news", body="Body",
                 mood="pos", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
                 item_type="news")
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    assert (cfg.output_root / "X5" / "news").exists()
    assert not (cfg.output_root / "X5" / "recommendations").exists()


def test_report_wipes_both_folders_on_regen(cfg: Config) -> None:
    """Reclassified item (news → recommendation) — старый MD в news/ удаляется."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    nid = _insert_news(conn, headline="Reclassify me", body="Body",
                      mood="pos", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
                      item_type="news")
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    news_md_v1 = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(news_md_v1) == 1

    # Reclassify: news → recommendation
    conn = db.connect(cfg.db_path)
    conn.execute("UPDATE news SET item_type='recommendation' WHERE id=?", (nid,))
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    news_md_v2 = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    rec_md_v2 = list((cfg.output_root / "X5" / "recommendations").rglob("*.md"))
    assert len(news_md_v2) == 0  # старый MD удалён
    assert len(rec_md_v2) == 1   # новый MD в recommendations/


def test_xlsx_has_two_sheets_with_recs_on_second(cfg: Config) -> None:
    """Excel workbook has two sheets: 'news' (только item_type='news') и
    'recommendations' (включает finam-recs из news.item_type='recommendation' +
    отдельную таблицу recommendations). Behavior change в задаче 06 — см.
    spec P2.5."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    # finam-style recommendation попадёт в news с item_type='recommendation'
    _insert_news(conn, headline="A", body="Body",
                 mood="pos", published_utc=datetime(2026, 5, 19, tzinfo=timezone.utc),
                 item_type="recommendation")
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    wb = load_workbook(cfg.output_root / "X5" / "news_list" / "data.xlsx")
    # Два листа в порядке news → recommendations
    assert wb.sheetnames == ["news", "recommendations"]

    # Sheet news — только заголовок, никаких finam-recs строк
    ws_news = wb["news"]
    news_headers = [ws_news.cell(1, c).value for c in range(1, ws_news.max_column + 1)]
    assert news_headers == ["date", "headline", "persons", "mood", "item_type"]
    assert ws_news.max_row == 1  # только header, нет данных

    # Sheet recommendations содержит finam-rec строку
    ws_recs = wb["recommendations"]
    rec_headers = [ws_recs.cell(1, c).value for c in range(1, ws_recs.max_column + 1)]
    assert rec_headers == [
        "date", "headline", "persons", "mood", "source",
        "target_price", "recommendation_action", "potential_pct", "multipliers",
    ]
    # finam-style рекомендация — target_price etc остаются NULL (legacy путь)
    assert ws_recs.max_row == 2
    assert ws_recs.cell(2, 2).value == "A"  # headline
    assert ws_recs.cell(2, 4).value == "pos"  # mood


# NOTE: sanitize_inline_code / clean_text покрыты в tests/test_text_cleanup.py.
# Здесь — только end-to-end через reporter (см. test_report_unescapes_stored_headlines).


def test_make_slug_unescapes_html_entities() -> None:
    """&quot;X5&quot; → x5, не quot_x5_quot."""
    s = reporter.make_slug("&quot;ИКС 5&quot; и &quot;Распадская&quot; получили")
    assert "quot" not in s
    assert s.startswith("икс_5")


def test_report_unescapes_stored_headlines(cfg: Config) -> None:
    """Headline/mood_reason с &quot; в БД — в файле и слаге уже без entity."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline='&quot;ИКС 5&quot; объявила результаты',
        body='Текст с &quot;цитатой&quot; внутри.',
        mood="pos",
        mood_reason='Позитив для бренда &quot;ИКС 5&quot;',
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    md_files = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(md_files) == 1
    name = md_files[0].name
    assert "quot" not in name, f"slug contains 'quot' entity token: {name}"
    text = md_files[0].read_text(encoding="utf-8")
    # Проверяем тело статьи (H1 + body), а не URL — в URL entity это легитимная часть.
    body_lines = [ln for ln in text.splitlines() if ln.startswith("# ") or ln == "Текст с \"цитатой\" внутри."]
    assert any('# "ИКС 5" объявила результаты' in ln for ln in body_lines), text
    assert "Текст с \"цитатой\" внутри." in text


def test_filename_uses_mood_reason_slug(cfg: Config) -> None:
    """Slug в имени файла строится из mood_reason, а не headline."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline="Очень длинный заголовок про X5 с массой подробностей и цифрами",
        body="Текст.",
        mood="pos",
        mood_reason="Сильный финансовый отчёт",
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    md_files = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(md_files) == 1
    name = md_files[0].name
    # mood_reason "Сильный финансовый отчёт" → "сильный_финансовый_отчет"
    assert "сильный_финансовый_отчет" in name
    # А части заголовка в слаге нет
    assert "очень_длинный" not in name


def test_filename_mood_reason_truncated_to_50_chars(cfg: Config) -> None:
    """Длинный mood_reason обрезается до 50 символов на границе слова."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    long_reason = "Позитивная новость про корпоративные результаты с очень-очень длинным текстом"
    _insert_news(
        conn,
        headline="X5",
        body="Body.",
        mood="pos",
        mood_reason=long_reason,
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    md_files = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(md_files) == 1
    name = md_files[0].stem  # без .md
    # Извлекаем slug часть: между датой+src и финальным _NN
    # Format: 2026_05_19_corp_<slug>_01
    parts = name.split("_")
    # date = parts[0..2], src=parts[3], slug=parts[4:-1], NN=parts[-1]
    slug_part = "_".join(parts[4:-1])
    assert len(slug_part) <= 50, f"slug {slug_part!r} length {len(slug_part)} > 50"


def test_filename_falls_back_to_headline_if_mood_reason_empty(cfg: Config) -> None:
    """Если mood_reason пуст — slug строится из headline."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline="Заголовок про X5",
        body="Body.",
        mood="neutral",
        mood_reason="",
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    md_files = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(md_files) == 1
    name = md_files[0].name
    assert "заголовок_про_x5" in name


def test_source_slug_known_codes() -> None:
    """x5_ir → corp; finam → fnm; rbc → rbc."""
    assert reporter._source_slug("x5_ir") == "corp"
    assert reporter._source_slug("finam") == "fnm"
    assert reporter._source_slug("rbc") == "rbc"


def test_source_slug_unknown_falls_back_to_first3_alnum() -> None:
    assert reporter._source_slug("interfax_news") == "int"
    assert reporter._source_slug("X") == "x"
    assert reporter._source_slug("") == "src"


def test_filename_includes_source_slug(cfg: Config) -> None:
    """В имени файла после даты идёт src-slug: 2026_05_19_corp_<slug>_NN.md."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline="X5 объявила результаты",
        body="Текст.",
        mood="pos",
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()
    reporter.report_all(cfg, "X5")
    md_files = list((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert len(md_files) == 1
    # _insert_news по дефолту использует source_id из x5_ir (см. fixture) → "corp"
    name = md_files[0].name
    assert name.startswith("2026_05_19_corp_"), f"unexpected name: {name}"
    assert name.endswith("_01.md")


def test_slug_basic() -> None:
    s = reporter.make_slug("Х5 объявляет о размере рекомендуемых дивидендов")
    assert s
    assert re.fullmatch(r"[а-яa-z0-9_]+", s)
    assert len(s) <= 60


def test_slug_first_five_words_only() -> None:
    s = reporter.make_slug("один два три четыре пять шесть семь восемь")
    # 6th word must not appear in slug
    assert "шесть" not in s
    assert "пять" in s


def test_slug_preserves_cyrillic_and_normalizes_yo() -> None:
    s = reporter.make_slug("Пятёрочка открыла магазин")
    assert s == "пятерочка_открыла_магазин"


def test_report_creates_md_persons_xlsx(cfg: Config) -> None:
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline="Шехтерман объявил план развития",
        body="Подробный текст пресс-релиза.",
        mood="pos",
        published_utc=datetime(2026, 5, 19, 7, 0, tzinfo=timezone.utc),
        persons_full_names=["Игорь Шехтерман"],
    )
    _insert_news(
        conn,
        headline="Наумова прокомментировала ситуацию",
        body="Ещё один пресс-релиз.",
        mood="neg",
        published_utc=datetime(2026, 5, 19, 8, 0, tzinfo=timezone.utc),
        persons_full_names=["Ольга Наумова"],
    )
    conn.commit()
    conn.close()

    results = reporter.report_all(cfg, "X5")
    assert len(results) == 1
    r = results[0]
    assert r.md_files == 2
    assert r.xlsx_rows == 2
    assert r.xlsx_written is True
    assert r.persons_rows == 2

    # Files exist
    out = cfg.output_root / "X5"
    md_files = sorted((out / "news" / "2026" / "2026_05").glob("*.md"))
    assert len(md_files) == 2
    # Filename starts with date and has a 2-digit ordinal suffix
    for p in md_files:
        assert p.name.startswith("2026_05_19_")
        assert p.stem.endswith(("_01", "_02"))

    # MD content: every file has frontmatter + headline + persons block
    moods_found: set[str] = set()
    for p in md_files:
        text = p.read_text(encoding="utf-8")
        assert text.startswith("---")
        assert "persons: [" in text
        for m in ("pos", "neg", "neutral"):
            if f"mood: {m}" in text:
                moods_found.add(m)
    assert moods_found == {"pos", "neg"}

    # persons.csv: aggregate freqs reflect the data we inserted
    with (out / "affiliate" / "persons.csv").open(encoding="utf-8") as fh:
        rows = {r["person"]: r for r in csv.DictReader(fh)}
    assert rows["Игорь Шехтерман"]["pos_freq"] == "1"
    assert rows["Игорь Шехтерман"]["total_freq"] == "1"
    assert rows["Ольга Наумова"]["neg_freq"] == "1"

    # xlsx: 1 header + 2 rows
    wb = load_workbook(out / "news_list" / "data.xlsx")
    ws = wb.active
    assert ws.cell(1, 1).value == "date"
    assert ws.max_row == 3


def test_timezone_conversion_to_moscow(cfg: Config) -> None:
    """A UTC publication just before midnight on the last day of a month
    must land in the *next* month's folder when converted to Moscow time
    (UTC+3)."""
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    # 2026-05-31 22:00 UTC == 2026-06-01 01:00 Moscow
    _insert_news(
        conn,
        headline="Новость на границе месяца",
        body="Тест таймзоны.",
        mood="neutral",
        published_utc=datetime(2026, 5, 31, 22, 0, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    files = list((cfg.output_root / "X5" / "news" / "2026" / "2026_06").glob("*.md"))
    assert len(files) == 1
    assert files[0].name.startswith("2026_06_01_")


def test_idempotent_regeneration(cfg: Config) -> None:
    db.init_db(cfg)
    conn = db.connect(cfg.db_path)
    _insert_news(
        conn,
        headline="Тест",
        body="Body",
        mood="pos",
        published_utc=datetime(2026, 5, 10, 9, tzinfo=timezone.utc),
    )
    conn.commit()
    conn.close()

    reporter.report_all(cfg, "X5")
    first = sorted((cfg.output_root / "X5" / "news").rglob("*.md"))
    reporter.report_all(cfg, "X5")
    second = sorted((cfg.output_root / "X5" / "news").rglob("*.md"))
    assert [p.name for p in first] == [p.name for p in second]


# Import here to keep top of file tidy
import re  # noqa: E402
